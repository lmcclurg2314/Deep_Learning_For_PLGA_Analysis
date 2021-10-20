##### Introduction
# This Python file takes the array that we organized in the previous 
# FCS_to_Array Python file to a .csv file for the logistic regression in R.

##### Loading modules
import pickle
import openpyxl

##### Loading data
allData = pickle.load(open("allData.obj", "rb"))
cytof_files = allData["cytof_files"]
expr_list = allData["expr_list"]
(_, n_cells, _, _) = expr_list.shape
marker_names = allData["marker_names"]
marker_names.remove("Time")

##### Preparing .csv file
wb = openpyxl.Workbook()
dest_filename = "data_unt_dox.csv"
ws1 = wb.active

##### Writing header into .csv file
for col in range(len(marker_names)):
    col = col + 1 # Starts at 1 and not 0
    _ = ws1.cell(column = col, row = 1, value = marker_names[col-1])
_ = ws1.cell(column = col + 1, row = 1, value = "dox_status")

##### Writing data into .csv file
for i in range(len(expr_list)):
    if (cytof_files["DOX_Ab"][i]): dox_status = 1
    else: dox_status = 0
    for j in range(len(expr_list[0])):
        for k in range(len(expr_list[0][0])):
            col = k + 1 # Starts at 1 and not 0
            row = i*n_cells + j + 2 # Also starts at 1 and 1 more for header
            # print(i)
            # print(j)
            # print(k)
            _ = ws1.cell(column = col, row = row, value = expr_list[i][j][k][0])
        _ = ws1.cell(column = col + 1, row = row, value = dox_status)

wb.save(filename = dest_filename)